import requests
import time
#TODO 测试结果杨幂微博 抓取数据365条 用时30分钟 一个进程16个线程
'''
#TODO 第二波测试
开始id  : 6124
开始时间： 2019-09-19 15:35:04
结束id ： 10968
结束时间：2019-09-19 19:22:18
爬取条数： 4844
用时 ： 3时47分
'''
# for i in range(20):
#     url="https://m.weibo.cn/api/container/getIndex?uid=5187664653&luicode=10000011&lfid=100103type%3D1%26q%3D%E9%82%93%E8%B6%85&type=uid&value=5187664653&containerid=1076035187664653&page="+str(i)
#     res=requests.get(url).json()
#     print(res["data"]["cardlistInfo"])
from urllib import parse
import re
def readxls():
    import openpyxl
    #print(parse.quote("新浪体育"))
    # 打开excel文件,获取工作簿对象
    wb = openpyxl.load_workbook(r'C:/Users/acer/Desktop/weibo.xlsx')
    s=requests.session()
    s.verify=False
    # 从表单中获取单元格的内容
    ws = wb.active  # 当前活跃的表单
    for i in range(10, 301):
        time.sleep(5)
        try:
            #model = Dynamicsource()
            # 获取A列的第一个对象
            name = ws['c' + str(i)].value
            print(name)
            parse_name=parse.quote(name)
            #GET https://m.weibo.cn/api/container/getIndex?containerid=100103type%3D3%26q%3D%E6%B2%88%E6%9C%88%E7%9A%84%E8%A1%A3%E6%A9%B1%E5%90%9B%26t%3D0&page_type=searchall
            res=s.get("https://m.weibo.cn/api/container/getIndex?containerid=100103type%3D3%26q%3D"+parse_name+"&page_type=searchall").json()
            data=res["data"]["cards"][1]["card_group"][0]["itemid"]
            print(data)
            names=re.search('q:(.*?)\|',data).group(1)
            if names==name:
                values=re.search('uid=(.*?)&',data).group(1)
                ws['e' + str(i)]=values
                print(ws['e' + str(i)])
                wb.save(r'C:/Users/acer/Desktop/weibo.xlsx')
        except Exception as e:
            print(e)
readxls()

