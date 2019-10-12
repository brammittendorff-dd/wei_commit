import requests
import re
import time
from lxml import etree
from pymongo import MongoClient
import datetime
from dateutil.relativedelta import relativedelta
class AiDou():

    def __init__(self):
        self.s=requests.session()
        self.s.headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36"
        }
        self.s.verify=False
                 #https://idol001.com/xingcheng/list/star-luhan-6621/2019/9/
        self.url="https://idol001.com/star/caixukun.html"
        conn = MongoClient('127.0.0.1', 27017)
        db = conn.mydb  # 连接mydb数据库，没有则自动创建
        self.my_set = db.caixukun # 使用didou集合，没有则自动创建
    def index_page(self):
        name=re.search(r'https://idol001.com/star/(.*?).html',self.url).group(1)
        res=self.s.get(self.url).text
        try:
            index_url=re.search('<a class="card-title card-route-head" href="(.*?)"',res).group(1)
            self.data_url(index_url,name)
        except:
            return
    def get_data_list(self):
        datetime_now = datetime.datetime.now()
        months = -1
        months_list = []
        while True:
            now_time = datetime_now - relativedelta(months=months)
            months += 1
            str_now_time = str(now_time).split(" ")[0]
            if str_now_time < '2016-01-01':
                break

            str_now_time=str_now_time.split("-")[0:2]
            str_now_time="/".join(str_now_time)
            months_list.append(str_now_time)
        return months_list
    def data_url(self,index_url,name):
        date_list=self.get_data_list()
        print(date_list)
        if not date_list:
            return
        for now_time in date_list:
            base_url=index_url+now_time
            self.parse_list(base_url,now_time,name)
    def parse_list(self,base_url,now_time,name):
        now_time=now_time.split("/")[0]+"年 "
        res=self.s.get(base_url)
        res.encoding="utf-8"
        html=etree.HTML(res.text)
        datas=html.xpath('.//table[@class="schedule-table"]//tr')
        if datas[1::]:

            for data in datas[1::]:
                item=self.xpath_parse(data,now_time,name)
                #print(item)
                if item:
                    print(item)
                    self.mongo_insert(item)
    def mongo_insert(self,item):
        #res = self.my_set.delete_one(item)
        res = self.my_set.find_one(item)
        if not res:
            self.my_set.insert_one(item)
    def xpath_parse(self,data,now_time,name):
        item={}
        try:
            item["name"]=name
            item["title"] = data.xpath('.//td[3]')[0].xpath('string(.)').replace("\n", "").strip()
            item["date"]=now_time+data.xpath('.//td[1]/text()')[0].replace("\n","").strip()
            item["detail_data"]=data.xpath('.//td[2]/text()')[0].replace("\n","").strip()
            item["type"]=data.xpath('.//td[4]/text()')[0].replace("\n","").strip()
            item["addr"]=data.xpath('.//td[5]/text()')[0].replace("\n","").strip()
            source=data.xpath('.//td[6]/text()')[0].replace("\n","").strip()
            if not source:
                sources=data.xpath('.//td[6]/a')
                source_list=[]
                for i in sources:
                    media_item={}
                    media_item["media"]=i.xpath('string(.)')
                    media_item["media_url"]=i.xpath('.//@href')[0]
                    source_list.append(media_item)
                item["source"] = source_list
            else:
                item["source"]=[]
            return item
        except:
            print("此条抓取失败。。。")
            return item

    #导入数据存入csv
    def daochu(self):
        res=self.my_set.find({"name":"xiaozhan"})
        data_list=[]
        for i in res:
            detail_list=[]
            for name,value in i.items():
                detail_list.append(str(value))
            data_list.append(detail_list)
        self.excel_save(data_list)

    def write_csv(self,data):
        print(data)
        import csv
        tmp = open("肖战.csv", 'a', newline='', encoding="utf-8")  # a表示在最后一行后面追加
        # newline以免出现写一行空一行
        # encoding 解决不能写入的错误

        csv_write = csv.writer(tmp)
        # csv_write.writerow(['id', 'eng_socre']) 写入列名
        csv_write.writerow(["id","姓名","活动","日期","详细日期","类型","地址","资源"])
        for item in data:

            csv_write.writerow(item)
        tmp.close()
    def excel_save(self,data):
        from openpyxl import Workbook
        wb = Workbook()  # 创建文件对象
        ws = wb.active  # 获取第一个sheet

        ws.title = "sheet"  # 设置sheet名称
        ws["A1"] = "姓名"
        ws["B1"] = "姓名"  # 设置表头（A1的位置）：为time
        ws["C1"] = "活动"  # 设置表头（B1的位置）：为year
        ws["D1"] = "日期"
        ws["E1"] = "详细日期"
        ws["F1"] = "类型"
        ws["G1"] = "地址"
        ws["H1"] = "资源"
        for i in data:

            ws.append(i)  # 添加第一行数据
        # 添加第二行数据
        wb.save("D:/肖战.xlsx")  # 保存



if __name__=="__main__":
    ad=AiDou()
    ad.index_page()
    #ad.daochu()


