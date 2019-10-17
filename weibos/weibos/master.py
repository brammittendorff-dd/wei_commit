import sys
sys.path.append('/home/commit/wei_commit/weibos')
import redis
from urllib import parse
import time
import datetime
from weibos.database.models import Dynamicsource
from weibos.database.db import session


class Master():

    def __init__(self):
        self.r = redis.Redis(host="47.110.95.150", port=6379, password="Bitgraph818")
        self.CELEBRITY_NEWS_API_URL = 'https://m.weibo.cn/api/container/getIndex?uid={}&luicode=10000011&lfid=100103type%3D1%26q%3D{}&type=uid&value={}&containerid={}&page={}'

    def weibo_redis(self):
        models = session.query(Dynamicsource).all()
        for model in models:
            if model.weibo_ID and model.id ==19:
                weiboid = parse.quote(model.name)
                weibo_burl = self.CELEBRITY_NEWS_API_URL.format(str(model.weibo_ID), weiboid,
                                                                str(model.weibo_ID), "107603"+str(model.weibo_ID), "0")
                print(weibo_burl)
                #self.r.rpush("weibo:start_urls", weibo_burl)
    def weibo_new_redis(self):
        models = session.query(Dynamicsource).all()
        for model in models:
            if model.weibo_ID:
                weiboid = parse.quote(model.name)
                weibo_burl = self.CELEBRITY_NEWS_API_URL.format(str(model.weibo_ID), weiboid,
                                                                str(model.weibo_ID), "107603"+str(model.weibo_ID), "0")
                #print(weibo_burl)
                self.r.rpush("weibo_new:start_urls", weibo_burl)

    # def weibo_test_redis(self):
    #
    #     weiboid = parse.quote("杨幂")
    #     weibo_burl = self.CELEBRITY_NEWS_API_URL.format(str(1195242865), weiboid,str(1195242865), "1076031195242865", "")
    #     self.r.rpush("weibo:start_urls", weibo_burl)
    def readxls(self):
        import openpyxl

        # 打开excel文件,获取工作簿对象
        wb = openpyxl.load_workbook(r'C:/Users/acer/Desktop/weibo.xlsx')
        # 从表单中获取单元格的内容
        ws = wb.active  # 当前活跃的表单
        for i in range(292,293):

      # 获取A列的第一个对象
            name=ws['c'+str(i)].value
            models=session.query(Dynamicsource).filter(Dynamicsource.name==name).all()
            if models:
                model=models[0]
                model.name=name
                model.weibo_ID=str(ws['d'+str(i)].value)
                print(model.name)
            else:
                model = Dynamicsource()
                model.name = name
                print(ws['d'+str(i)].value)
                model.weibo_ID=str(ws['d'+str(i)].value)

            session.add(model)
            try:
                session.commit()
            except:
                print(1111)
                raise Exception
ms = Master()
#ms.readxls()
#ms.weibo_redis()
ms.weibo_new_redis()
#ms.weibo_test_redis()


